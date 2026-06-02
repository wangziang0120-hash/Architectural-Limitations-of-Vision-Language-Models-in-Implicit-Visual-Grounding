#!/usr/bin/env python3
"""生成Excel对比表格 - 支持多模型"""
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from config import BASE_DIR, METHODS, MODELS

METHOD_COLORS = {
    "baseline": "FF4444",
    "prompt": "4444FF",
    "blur": "FF8800",
    "blackout": "8800FF",
    "crop": "00AA00",
}


def load_jsonl(path: Path) -> list[dict]:
    data = []
    if path.exists():
        with path.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    data.append(json.loads(line))
    return data


def create_excel(model_key: str):
    model_cfg = MODELS[model_key]
    results_dir = Path(BASE_DIR) / model_cfg["output_dir"]
    output_file = results_dir / f"evaluation_results_{model_key}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    summary_sheet = wb.create_sheet("Summary")
    headers = ["Method", "Total", "Correct", "Acc@0.5", "Avg IoU"]
    for col, header in enumerate(headers, 1):
        cell = summary_sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="333333")
        cell.alignment = Alignment(horizontal="center")

    for row_idx, (method, method_cfg) in enumerate(METHODS.items(), start=2):
        results = load_jsonl(results_dir / method_cfg["results_file"])
        correct = sum(1 for r in results if r.get("correct"))
        total = len(results)
        iou_vals = [r.get("iou", 0) for r in results]
        avg_iou = sum(iou_vals) / len(iou_vals) if iou_vals else 0
        acc = correct / total if total > 0 else 0

        summary_sheet.cell(row=row_idx, column=1, value=method).font = Font(
            bold=True, color=METHOD_COLORS.get(method)
        )
        summary_sheet.cell(row=row_idx, column=2, value=total)
        summary_sheet.cell(row=row_idx, column=3, value=correct)
        summary_sheet.cell(row=row_idx, column=4, value=acc).number_format = "0.0%"
        summary_sheet.cell(row=row_idx, column=5, value=avg_iou).number_format = "0.0000"

        detail_sheet = wb.create_sheet(method[:31])
        detail_headers = [
            "#", "Image ID", "Question", "GT BBox", "Pred BBox",
            "IoU", "Correct", "Status",
        ]
        for col_idx, header in enumerate(detail_headers, 1):
            cell = detail_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=METHOD_COLORS.get(method, "333333"))
            cell.alignment = Alignment(horizontal="center")

        for i, r in enumerate(results):
            row = i + 2
            detail_sheet.cell(row=row, column=1, value=i + 1)
            detail_sheet.cell(row=row, column=2, value=r.get("image_id", ""))
            detail_sheet.cell(row=row, column=3, value=r.get("question", "")[:80])

            gt = r.get("gt_bbox")
            if gt:
                detail_sheet.cell(
                    row=row, column=4,
                    value=f"[{gt[0]:.1f}, {gt[1]:.1f}, {gt[2]:.1f}, {gt[3]:.1f}]",
                )

            pred = r.get("pred_bbox")
            if pred:
                detail_sheet.cell(
                    row=row, column=5,
                    value=f"[{pred[0]:.1f}, {pred[1]:.1f}, {pred[2]:.1f}, {pred[3]:.1f}]",
                )

            detail_sheet.cell(
                row=row, column=6, value=r.get("iou", 0)
            ).number_format = "0.0000"

            correct_cell = detail_sheet.cell(
                row=row, column=7, value="Yes" if r.get("correct") else "No"
            )
            correct_cell.font = Font(
                color="008800" if r.get("correct") else "CC0000", bold=True
            )

            detail_sheet.cell(row=row, column=8, value=r.get("status", ""))

        for col_letter, width in [
            ("A", 5), ("B", 35), ("C", 50), ("D", 25),
            ("E", 25), ("F", 10), ("G", 10), ("H", 15),
        ]:
            detail_sheet.column_dimensions[col_letter].width = width

    summary_sheet.column_dimensions["A"].width = 20
    summary_sheet.column_dimensions["B"].width = 10
    summary_sheet.column_dimensions["C"].width = 10
    summary_sheet.column_dimensions["D"].width = 12
    summary_sheet.column_dimensions["E"].width = 12

    wb.save(output_file)
    print(f"Excel已保存: {output_file}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(f"用法: python {sys.argv[0]} <model_key>")
        print(f"可用模型: {', '.join(MODELS.keys())}")
        sys.exit(1)
    create_excel(sys.argv[1])
